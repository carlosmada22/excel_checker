# -*- coding: utf-8 -*-
"""
Created on Thu Oct 10

@author: cmadaria
"""

import re
import openpyxl
import logging
import csv
import io
import base64
from datetime import datetime
from cryptography.fernet import Fernet, InvalidToken
from django.conf import settings

logger = logging.getLogger('myapp')

# Instantiate the Fernet class with the secret key
cipher_suite = Fernet(settings.SECRET_ENCRYPTION_KEY)

# Encrypt the password
def encrypt_password(plain_text_password):
    encrypted_password = cipher_suite.encrypt(plain_text_password.encode('utf-8'))
    return encrypted_password.decode('utf-8')  # Return as a string

def decrypt_password(encrypted_password):
    try:
        # Remove the manual padding correction, Fernet handles it automatically
        decrypted_password = cipher_suite.decrypt(encrypted_password.encode('utf-8'))
        return decrypted_password.decode('utf-8')
    except InvalidToken as e:
        print(f"Decryption failed: {str(e)}")
        raise InvalidToken("Decryption failed due to an invalid token.")
    except Exception as e:
        print(f"Decryption error: {str(e)}")
        raise e


def name_checker(file_name):
    # Define the pattern for a valid file name
    pattern = r"^(collection_type|object_type|dataset_type|vocabulary)_([\w.]+)_(v\d+)_([a-zA-Z0-9]+(?:\.[0-9]+)?)_([a-zA-Z0-9]+)\.(xls|xlsx)$"
 
    # Check if the file name matches the pattern
    match = re.match(pattern, file_name)

    # Return specific errors and positions
    errors = []
    file_name = file_name.split(".xls")
    file_parts = file_name[0].split("_")
    if len(file_parts) < 5:
        errors.append("⦿ <strong>Invalid name format</strong>. The name should contain different fields separated by underscores (_). Consult the wiki to see which ones.")
        return ["\n".join(errors), file_name, False]
    creator = file_parts.pop(-1)
    section = file_parts.pop(-1)
    version = file_parts.pop(-1)
    etype = file_parts.pop(0)
    if (etype == "object" or etype == "collection" or etype == "dataset"):
        etype = etype + "_" + file_parts.pop(0)
    code = "_".join(file_parts)

    if match:
        # Extract parts of the file name
        entity_type, entity_name, version, division, contact_person, extension = match.groups()
        print(entity_type, entity_name, version, division, contact_person, extension)
        return ["⦿ <strong>File name: OK!</strong>", code, True]
    else:   
        if not re.match(r"^(collection_type|object_type|dataset_type|vocabulary)$", etype):
            errors.append("⦿ <strong>Invalid entity type</strong> at position 1.")
        if not re.match(r"^([\w.]+)$", code):
            errors.append("⦿ <strong>Invalid entity name</strong> at position 2.")
        if not re.match(r"^(v\d+)$", version):
            errors.append("⦿ <strong>Invalid version</strong> at position 3.")
        if not re.match(r"^([a-zA-Z0-9]+(?:\.[0-9]+)?)$", section):
            errors.append("⦿ <strong>Invalid division</strong> at position 4.")
        if not re.match(r"^[a-zA-Z0-9]+$", creator):
            errors.append("⦿ <strong>Invalid contact person</strong> at position 5.")
            
        return ["\n".join(errors), code, False]
    

def index_to_excel_column(index):
    column = ''
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        column = chr(65 + remainder) + column
    return column


def check_properties(sheet, errors):
    expected_terms = [
        "Version",
        "Code",
        "Description",
        "Mandatory",
        "Show in edit views",
        "Section",
        "Property label",
        "Data type",
        "Vocabulary code"
    ]
    row_headers = [cell.value for cell in sheet[4]]
    for term in expected_terms:
        if (term not in row_headers):
            if term in ("Mandatory","Show in edit views","Section"):
                errors.append(f"⦿ <em>Warning</em>: '{term}' not found in the properties headers.")
            else:
                errors.append(f"⦿ <strong>Error</strong>: '{term}' not found in the properties headers.")
        else:
             # Find the index of the term in the second row
             term_index = row_headers.index(term) + 1
             term_letter = index_to_excel_column(term_index)
             #print(term_index)
             
             # Check the column below "Version"
             if term == "Version":
                 column_below_version = []
                 for cell in sheet[term_letter][4:]:
                     if cell.value is not None:
                         column_below_version.append(cell.value)
                     else:
                         pass

                 # Check if any value in the column is not an integer
                 non_integer_indices = [i + 5 for i, cell in enumerate(column_below_version) if not (str(cell).isnumeric() or "$" in str(cell))]
                 if non_integer_indices:
                     # Append an error indicating the positions (row numbers) that are not integers
                     errors.append(f"<strong>Error</strong>: Values not valid found in the 'Version' column (they should be Integers) at row(s): {', '.join(map(str, non_integer_indices))}")

            # Check the column below "Code"
             elif term == "Code":
                column_below_code = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_code.append(cell.value)
                    else:
                        pass
                invalid_codes = [i + 5 for i, cell in enumerate(column_below_code) if not (re.match(r'^\$?[A-Z0-9_.]+$', str(cell)) or "$" in str(cell))]
                if invalid_codes:
                    # Append an error indicating the positions (row numbers) with invalid values for the current term
                    errors.append(f"<strong>Error</strong>: Invalid code found in the '{term}' column at row(s): {', '.join(map(str, invalid_codes))}")
                    
                #check that all the properties of the object are different using a set (unique terms):
                if len(set(column_below_code)) != len(column_below_code):
                    seen_props = set()
                    repeated_props = set()
                    for prop in column_below_code:
                        if prop in seen_props:
                            repeated_props.add(prop)
                        else:
                            seen_props.add(prop)
                    errors.append(f"<strong>Error</strong>: The following properties are repeated: {repeated_props}. Please, delete the duplicates, and leave just one occurence")

            
            
            # Check the cell below "Description"
             elif term == "Description":
                column_below_description = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_description.append(cell.value)
                    else:
                        pass
                invalid_indices = [i + 5 for i, cell in enumerate(column_below_description) if not (re.match(r'.*//.*', str(cell)) or "$" in str(cell))]
                if invalid_indices:
                    errors.append(f"<strong>Error</strong>: Invalid value(s) found in the '{term}' column at row(s): {', '.join(map(str, invalid_indices))}. Description should follow the schema: English Description + '//' + German Description.")

            # Check the cell below "Mandatory"
             elif term == "Mandatory":
                column_below_mandatory = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_mandatory.append(str(cell.value).upper())
                    else:
                        pass
                invalid_mandatory = [i + 5 for i, cell in enumerate(column_below_mandatory) if (cell not in ["TRUE", "FALSE"] and "$" not in str(cell))]
                if invalid_mandatory:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_mandatory))}. Accepted values: TRUE, FALSE")

            # Check the cell below "Show in edit views"
             elif term == "Show in edit views":
                column_below_show = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_show.append(str(cell.value).upper())
                    else:
                        pass
                invalid_show = [i + 5 for i, cell in enumerate(column_below_show) if (cell not in ["TRUE", "FALSE"] and "$" not in str(cell))]
                if invalid_show:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_show))}. Accepted values: TRUE, FALSE")

            # Check the cell below "Section"
             elif term == "Section":
                column_below_section = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_section.append(cell.value) if '$' not in cell.value else column_below_section.append(cell.value.replace('$', ''))
                    else:
                        pass
                    
                invalid_section = [i + 5 for i, cell in enumerate(column_below_section) if not (re.match(r'^[A-Z][A-Za-z]*(?:\s[A-Z][A-Za-z]*)*$', str(cell)) or "$" in str(cell))]
                if invalid_section:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_section))}. Each word in the Section should start with a capital letter.")
            
                # Group Check: Ensure all properties within the same section are grouped together
                seen_sections = {}
                non_contiguous_rows = []
                
                for i, current_value in enumerate(column_below_section):
                    if current_value in seen_sections:
                        # If the value has been seen before but the row is not contiguous, add an error
                        if seen_sections[current_value] != i - 1:
                            non_contiguous_rows.append(i + 5)
                    seen_sections[current_value] = i  # Update the last seen row index for the current value
                
                if non_contiguous_rows:
                    errors.append(f"<strong>Error</strong>: Non-contiguous rows found for the same 'Section' value at row(s): {', '.join(map(str, non_contiguous_rows))}. Ensure that all properties within the same Section are grouped together.")
            
                # Predefined section order (fixed order)
                predefined_section_order = ["General Information", "Additional Information", "Comments"]
            
                # Validate contiguous groups and predefined section order
                seen_sections = set()
                previous_section_type = None
                section_errors = []  # Store section-specific errors
                additional_info_seen = False  # Flag to track if "Additional Information" has been seen
                comments_seen = False  # Flag to track if "Comments" has been seen
            
                # Traverse the section list
                for i, section in enumerate(column_below_section):
                    if section in predefined_section_order:
                        if section == "General Information":
                            if previous_section_type not in [None, "General Information"]:
                                section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'General Information' should only appear at the beginning.")
                        elif section == "Additional Information":
                            if previous_section_type not in ["General Information", "user-defined"]:
                                section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'Additional Information' should appear after 'General Information' and any user-defined sections.")
                            additional_info_seen = True  # Mark that "Additional Information" has been encountered
                        elif section == "Comments":
                            if previous_section_type not in ["General Information", "user-defined", "Additional Information"]:
                                section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'Comments' should appear after 'Additional Information'.")
                            comments_seen = True  # Mark that "Comments" has been encountered
                        previous_section_type = section
                    else:
                        # User-defined section
                        if comments_seen:
                            section_errors.append(f"<strong>Error</strong> at row {i + 5}: User-defined section '{section}' cannot appear after 'Comments'.")
                        if additional_info_seen and not comments_seen:
                            section_errors.append(f"<strong>Error</strong> at row {i + 5}: User-defined section '{section}' cannot appear after 'Additional Information' but before 'Comments'.")
                        previous_section_type = "user-defined"
            
                # Output any errors
                if section_errors:
                    for error in section_errors:
                        errors.append(error)
            
            # Check the cell below "Property label"
             elif term == "Property label":
                column_below_label = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_label.append(cell.value)
                    else:
                        pass
                invalid_label = [i + 5 for i, cell in enumerate(column_below_label) if not (re.match(r'.*', str(cell)) or "$" in str(cell))]
                if invalid_label:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_label))}. Specify the property label as text format")
                    
                # Dynamically find the "Section" column
                if "Section" in row_headers:
                    section_index = row_headers.index("Section") + 1
                    section_letter = index_to_excel_column(section_index)
                    column_below_section = [cell.value for cell in sheet[section_letter][4:]]

                    # New check: "Notes" in "Property label" should correspond to "Additional Information" in "Section"
                    for i, label_value in enumerate(column_below_label):
                        if label_value == "Notes":
                            section_value = column_below_section[i]
                            if section_value != "Additional Information":
                                errors.append(f"<strong>Error</strong>: 'Notes' found in the 'Property label' column at row {i + 5}, but corresponding 'Section' column does not contain 'Additional Information'. Value found: {section_value}")


            # Check the cell below "Data type"
             elif term == "Data type":
                column_below_type = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_type.append(str(cell.value).upper())
                    else:
                        pass
                invalid_type = [i + 5 for i, cell in enumerate(column_below_type) if (cell not in ["INTEGER", "REAL", "VARCHAR", "MULTILINE_VARCHAR", "HYPERLINK", "BOOLEAN", "CONTROLLEDVOCABULARY", "XML", "TIMESTAMP", "DATE", "SAMPLE"] and "$" not in str(cell))]
                if invalid_type:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_type))}. Accepted types: INTEGER, REAL, VARCHAR, MULTILINE_VARCHAR, HYPERLINK, BOOLEAN, CONTROLLEDVOCABULARY, XML, TIMESTAMP, DATE, SAMPLE")

            # Check the column below "Vocabulary code"
             elif term == "Vocabulary code":
                column_below_vocab = sheet[term_letter][4:]
                invalid_vocab = [i + 5 for i, cell in enumerate(column_below_vocab) if cell.value and not (re.match(r'^\$?[A-Z0-9_.]', str(cell.value)) or "$" not in str(cell))]
                if invalid_vocab:
                    # Append an error indicating the positions (row numbers) with invalid values for the current term
                    errors.append(f"<strong>Error</strong>: Invalid vocabulary code found in the '{term}' column at row(s): {', '.join(map(str, invalid_vocab))}")
    
    return errors

def check_vocab_terms(sheet, errors):
    expected_terms = [
        "Version",
        "Code",
        "Label"
        "Description"
    ]
    row_headers = [cell.value for cell in sheet[4]]
    for term in expected_terms:
        if term not in row_headers:
            errors.append(f"<strong>Error</strong>: '{term}' not found in the vocabulary term headers.")
        else:
             # Find the index of the term in the second row
             term_index = row_headers.index(term) + 1
             term_letter = index_to_excel_column(term_index)
             #print(term_index)
             
             # Check the column below "Version"
             if term == "Version":
                 column_below_version = []
                 for cell in sheet[term_letter][4:]:
                     if cell.value is not None:
                         column_below_version.append(cell.value)
                     else:
                         pass

                 # Check if any value in the column is not an integer
                 non_integer_indices = [i + 5 for i, cell in enumerate(column_below_version) if not str(cell).isnumeric()]
                 if non_integer_indices:
                     # Append an error indicating the positions (row numbers) that are not integers
                     errors.append(f"<strong>Error</strong>: Values not valid found in the 'Version' column (they should be Integers) at row(s): {', '.join(map(str, non_integer_indices))}")

            # Check the column below "Code"
             elif term == "Code":
                column_below_code = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_code.append(cell.value)
                    else:
                        pass
                invalid_codes = [i + 5 for i, cell in enumerate(column_below_code) if not re.match(r'^\$?[A-Z0-9_.]+$', str(cell))]
                if invalid_codes:
                    # Append an error indicating the positions (row numbers) with invalid values for the current term
                    errors.append(f"<strong>Error</strong>: Invalid code found in the '{term}' column at row(s): {', '.join(map(str, invalid_codes))}")
                
                #check that all the properties of the object are different using a set (unique terms):
                if len(set(column_below_code)) != len(column_below_code):
                    seen_terms = set()
                    repeated_terms = set()
                    for term in column_below_code:
                        if term in seen_terms:
                            repeated_terms.add(term)
                        else:
                            seen_terms.add(term)
                    errors.append(f"<strong>Error</strong>: The following vocabulary terms are repeated: {repeated_terms}. Please, delete the duplicates, and leave just one occurence")

            
            
            # Check the cell below "Description"
             elif term == "Description":
                column_below_description = sheet[term_letter][4:]
                invalid_description = [i + 5 for i, cell in enumerate(column_below_description) if cell.value and not re.match(r'.*//.*', str(cell.value))]
                if invalid_description:
                    errors.append(f"<strong>Error</strong>: Invalid value(s) found in the '{term}' column at row(s): {', '.join(map(str, invalid_description))}. Description should follow the schema: English Description + '//' + German Description.")

            # Check the cell below "Mandatory"
             elif term == "Label":
                column_below_label = sheet[term_letter][4:]
                invalid_label = [i + 5 for i, cell in enumerate(column_below_label) if cell.value and not re.match(r'.*', str(cell.value))]
                if invalid_label:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_label))}. Specify the label as text format")
            
    return "\n".join(errors)

def content_checker(file_name, name_ok):
    logger.info(f"Type {type(file_name)} of file {file_name}")
    workbook = openpyxl.load_workbook(file_name)
    errors = []  
    
    if(name_ok):
        file_name = file_name.name.split(".xls")
        file_parts = file_name[0].split("_")
        file_parts.pop(-1)
        file_parts.pop(-1)
        version = file_parts.pop(-1)
        etype = file_parts.pop(0)
        if (etype == "object" or etype == "collection" or etype == "dataset"):
            etype = etype + "_" + file_parts.pop(0)
        code = "_".join(file_parts)
    else:
        version, etype, code = "", "", ""

    sheet = workbook.active
    
    filtered_rows = []
    
    for row in sheet.iter_rows(min_row=1, values_only=True):
    # Check if any cell in the row contains "$"
        if any("$" in str(cell) for cell in row):
            filtered_rows.append(["$" + str(cell) if cell is not None else None for cell in row])
        else:
            # If the row passed the check, add it to the filtered list
            filtered_rows.append(row)
    
    #remove all the rows in the sheet
    sheet.delete_rows(0, sheet.max_row)

    # Append the filtered rows to the sheet
    for row_data in filtered_rows:
        sheet.append(row_data)

    # Access a specific cell (e.g., cell A1)
    cell_value_A1 = sheet['A1'].value
    print(f"Entity Type: {cell_value_A1}")
    
    entity_types = ["SAMPLE_TYPE", "EXPERIMENT_TYPE", "DATASET_TYPE", "PROPERTY_TYPE", "VOCABULARY_TYPE"]
    if cell_value_A1 not in entity_types:
        errors.append("⦿ The entity type (cell A1) should be one of the following: SAMPLE_TYPE, EXPERIMENT_TYPE, DATASET_TYPE, PROPERTY_TYPE, VOCABULARY_TYPE")
        return "".join(errors)
    else:
        if cell_value_A1 == "SAMPLE_TYPE":
            expected_terms = [
                "Version",
                "Code",
                "Description",
                "Validation script",
                "Generated code prefix",
                "Auto generate codes",
            ]
            second_row_values = [cell.value for cell in sheet[2]]
            for term in expected_terms:
                if term not in second_row_values:
                    errors.append(f"<strong>Error</strong>: '{term}' not found in the entity headers.")
                else:
                     # Find the index of the term in the second row
                     term_index = second_row_values.index(term)

                     # Check the cell below "Version"
                     if term == "Version":
                        cell_below_version = sheet.cell(row=3, column=term_index + 1)
                        if str(cell_below_version.value) != version[1:]:
                            errors.append("<strong>Error</strong>: The version should be the same one indicated in the file name")

                    # Check the cell below "Code"
                     elif term == "Code":
                        cell_below_code = sheet.cell(row=3, column=term_index + 1)
                        if cell_below_code.value != code:
                            errors.append("⦿ <strong>Error</strong>: The code should be the same one indicated in the file name")
                    
                    
                    # Check the cell below "Description"
                     elif term == "Description":
                        cell_below_description = sheet.cell(row=3, column=term_index + 1)
                        description_pattern = re.compile(r".*//.*")
                        if not description_pattern.match(cell_below_description.value):
                            errors.append("<strong>Error</strong>: Description should follow the schema: English Description + '//' + German Description.")

                    # Check the cell below "Generated code prefix"
                     elif term == "Generated code prefix":
                        cell_below_generated_code = sheet.cell(row=3, column=term_index + 1)
                        code_replace = code.replace('_', '.').split('.')
                        ext_code = [word[:3].upper() for word in code_replace]
                        generated_code = '.'.join(ext_code)
                        if cell_below_generated_code.value != generated_code:
                            errors.append("<em>Warning</em>: It is recommended that the value of 'Generated code prefix' be the first three letters of each part of the 'Code' separated by dots ['.'].")

                    # Check the cell below "Validation script"
                     elif term == "Validation script":
                        cell_below_validation = sheet.cell(row=3, column=term_index + 1)
                        validation_pattern = re.compile(r"^[A-Za-z0-9_]+\.py$")
                        if cell_below_validation.value and not validation_pattern.match(cell_below_validation.value):
                             errors.append("<strong>Error</strong>: Validation script should follow the schema: Words and/or numbers separated by '_' and ending in '.py'")


                    # Check the cell below "Auto generate codes"
                     elif term == "Auto generate codes":
                        cell_below_auto_generate = sheet.cell(row=3, column=term_index + 1)
                        auto_code = cell_below_auto_generate.value
                        if (auto_code == True): auto_code = "TRUE"
                        if (auto_code == False): auto_code = "FALSE"
                        if auto_code not in ["TRUE", "FALSE"]:
                            errors.append("<strong>Error</strong>: Value below 'Auto generate codes' should be 'TRUE' or 'FALSE'.")
            
            errors = check_properties(sheet, errors)      
            
        elif cell_value_A1 == "EXPERIMENT_TYPE" or cell_value_A1 == "DATASET_TYPE":
            expected_terms = [
                "Version",
                "Code",
                "Description",
                "Validation script"
            ]
            second_row_values = [cell.value for cell in sheet[2]]
            for term in expected_terms:
                if term not in second_row_values:
                    errors.append(f"<strong>Error</strong>: '{term}' not found in the second row.")
                else:
                     # Find the index of the term in the second row
                     term_index = second_row_values.index(term)

                     # Check the cell below "Version"
                     if term == "Version":
                        cell_below_version = sheet.cell(row=3, column=term_index + 1)
                        if str(cell_below_version.value) != version[1:]:
                            errors.append("<strong>Error</strong>: The version should be the same one indicated in the file name")

                    # Check the cell below "Code"
                     elif term == "Code":
                        cell_below_code = sheet.cell(row=3, column=term_index + 1)
                        if cell_below_code.value != code:
                            errors.append("<strong>Error</strong>: The code should be the same one indicated in the file name")
                    
                    
                    # Check the cell below "Description"
                     elif term == "Description":
                        cell_below_description = sheet.cell(row=3, column=term_index + 1)
                        description_pattern = re.compile(r".*//.*")
                        if not description_pattern.match(cell_below_description.value):
                            errors.append("<strong>Error</strong>: Description should follow the schema: English Description + '//' + German Description.")
            
            
                    # Check the cell below "Validation script"
                     elif term == "Validation script":
                        cell_below_validation = sheet.cell(row=3, column=term_index + 1)
                        validation_pattern = re.compile(r"^[A-Za-z0-9_]+\.py$")
                        if cell_below_validation.value and not validation_pattern.match(cell_below_validation.value):
                            errors.append("<strong>Error</strong>: Validation script should follow the schema: Words and/or numbers separated by '_' and ending in '.py'")

            errors = check_properties(sheet, errors) 
            
        elif cell_value_A1 == "VOCABULARY_TYPE":
            expected_terms = [
                "Version",
                "Code",
                "Description"
            ]
            second_row_values = [cell.value for cell in sheet[2]]
            for term in expected_terms:
                if term not in second_row_values:
                    errors.append(f"<strong>Error</strong>: '{term}' not found in the second row.")
                else:
                     # Find the index of the term in the second row
                     term_index = second_row_values.index(term)

                     # Check the cell below "Version"
                     if term == "Version":
                        cell_below_version = sheet.cell(row=3, column=term_index + 1)
                        if str(cell_below_version.value) != version[1:]:
                            errors.append("<strong>Error</strong>: The version should be the same one indicated in the file name. Value found: {cell_below_version.value}")

                    # Check the cell below "Code"
                     elif term == "Code":
                        cell_below_code = sheet.cell(row=3, column=term_index + 1)
                        if cell_below_code.value != code:
                            errors.append("<strong>Error</strong>: The code should be the same one indicated in the file name. Value found: {cell_below_code.value}")
                    
                    
                    # Check the cell below "Description"
                     elif term == "Description":
                        cell_below_description = sheet.cell(row=3, column=term_index + 1)
                        description_pattern = re.compile(r".*//.*")
                        if not description_pattern.match(cell_below_description.value):
                            errors.append("<strong>Error</strong>: Description should follow the schema: English Description + '//' + German Description. Value found: {cell_below_description.value}")
            
            errors = check_vocab_terms(sheet, errors)

        elif cell_value_A1 == "PROPERTY_TYPE":
            expected_terms = [
                "Version",
                "Code",
                "Description",
                "Mandatory",
                "Show in edit views",
                "Section",
                "Property label",
                "Data type",
                "Vocabulary code"
            ]
            second_row_values = [cell.value for cell in sheet[2]]
            for term in expected_terms:
                if term not in second_row_values:
                    errors.append(f"<strong>Error</strong>: '{term}' not found in the second row.")
                else:
                     # Find the index of the term in the second row
                     term_index = second_row_values.index(term) + 1


                     # Check the column below "Version"
                     if term == "Version":
                        column_below_version = sheet[term_index][2:]
                        # Check if any value in the column is not an integer
                        non_integer_cells = [(i + 3, cell.value) for i, cell in enumerate(column_below_version) if not isinstance(cell.value, int)]
                        if non_integer_cells:
                            # Append an error indicating the positions (row numbers) that are not integers
                            non_integer_indices = [str(row) for row, _ in non_integer_cells]
                            invalid_values = [str(value) for _, value in non_integer_cells]
                            errors.append(f"<strong>Error</strong>: Values not valid found in the 'Version' column (they should be Integers) at row(s): {', '.join(non_integer_indices)}. Value(s) found: {', '.join(invalid_values)}")

                    # Check the column below "Code"
                     elif term == "Code":
                        column_below_code = sheet[term_index][2:]
                        invalid_codes = [(i + 3, cell.value) for i, cell in enumerate(column_below_code) if not re.match(r'^\$?[A-Z0-9_.]+$', str(cell.value))]
                        if invalid_codes:
                            invalid_rows = [str(row) for row, _ in invalid_codes]
                            invalid_values = [str(value) for _, value in invalid_codes]
                            errors.append(f"<strong>Error</strong>: Invalid code found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Value(s) found: {', '.join(invalid_values)}")
                    
                    # Check the cell below "Description"
                     elif term == "Description":
                        column_below_description = sheet[term_index][2:]
                        invalid_descriptions = [(i + 3, cell.value) for i, cell in enumerate(column_below_description) if not re.match(r'.*//.*', str(cell.value))]
                        if invalid_descriptions:
                            invalid_rows = [str(row) for row, _ in invalid_descriptions]
                            invalid_values = [str(value) for _, value in invalid_descriptions]
                            errors.append(f"<strong>Error</strong>: Invalid value(s) found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Description should follow the schema: English Description + '//' + German Description. Value(s) found: {', '.join(invalid_values)}")
                    
                    # Check the cell below "Mandatory"
                     elif term == "Mandatory":
                        column_below_mandatory = sheet[term_index][2:]
                        invalid_mandatory = [(i + 3, cell.value) for i, cell in enumerate(column_below_mandatory) if cell.value not in ["TRUE", "FALSE"]]
                        if invalid_mandatory:
                            invalid_rows = [str(row) for row, _ in invalid_mandatory]
                            invalid_values = [str(value) for _, value in invalid_mandatory]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Accepted values: TRUE, FALSE. Value(s) found: {', '.join(invalid_values)}")
                    
                    # Check the cell below "Show in edit views"
                     elif term == "Show in edit views":
                        column_below_show = sheet[term_index][2:]
                        invalid_show = [(i + 3, cell.value) for i, cell in enumerate(column_below_show) if cell.value not in ["TRUE", "FALSE"]]
                        if invalid_show:
                            invalid_rows = [str(row) for row, _ in invalid_show]
                            invalid_values = [str(value) for _, value in invalid_show]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Accepted values: TRUE, FALSE. Value(s) found: {', '.join(invalid_values)}")
                    
                     elif term == "Section":
                        column_below_section = sheet[term_index][2:]
                        print(column_below_section)
                        invalid_section = [(i + 3, cell.value) for i, cell in enumerate(column_below_section) if not re.match(r'^[A-Z][a-z]*(?:\s[A-Z][a-z]*)*$', str(cell.value))]
                        if invalid_section:
                            invalid_rows = [str(row) for row, _ in invalid_section]
                            invalid_values = [str(value) for _, value in invalid_section]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Each word should start with a capital letter. Value(s) found: {', '.join(invalid_values)}")
                    
                        seen_sections = {}
                        non_contiguous_rows = []

                        for i, current_value in enumerate(column_below_section):
                            if current_value in seen_sections:
                                # If the value has been seen before but the row is not contiguous, add an error
                                if seen_sections[current_value] != i - 1:
                                    non_contiguous_rows.append(i + 5)
                            seen_sections[current_value] = i  # Update the last seen row index for the current value

                        if non_contiguous_rows:
                            errors.append(f"<strong>Error</strong>: Non-contiguous rows found for the same 'Section' value at row(s): {', '.join(map(str, non_contiguous_rows))}. Ensure that all properties within the same Section are grouped together.")
                            
                        # Predefined section order (fixed order)
                        predefined_section_order = ["General Information", "Additional Information", "Comments"]
                    
                        # Validate contiguous groups and predefined section order
                        seen_sections = set()
                        previous_section_type = None
                        section_errors = []  # Store section-specific errors
                        additional_info_seen = False  # Flag to track if "Additional Information" has been seen
                        comments_seen = False  # Flag to track if "Comments" has been seen
                    
                        # Traverse the section list
                        for i, section in enumerate(column_below_section):
                            if section in predefined_section_order:
                                if section == "General Information":
                                    if previous_section_type not in [None, "General Information"]:
                                        section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'General Information' should only appear at the beginning.")
                                elif section == "Additional Information":
                                    if previous_section_type not in ["General Information", "user-defined"]:
                                        section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'Additional Information' should appear after 'General Information' and any user-defined sections.")
                                    additional_info_seen = True  # Mark that "Additional Information" has been encountered
                                elif section == "Comments":
                                    if previous_section_type not in ["General Information", "user-defined", "Additional Information"]:
                                        section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'Comments' should appear after 'Additional Information'.")
                                    comments_seen = True  # Mark that "Comments" has been encountered
                                previous_section_type = section
                            else:
                                # User-defined section
                                if comments_seen:
                                    section_errors.append(f"<strong>Error</strong> at row {i + 5}: User-defined section '{section}' cannot appear after 'Comments'.")
                                if additional_info_seen and not comments_seen:
                                    section_errors.append(f"<strong>Error</strong> at row {i + 5}: User-defined section '{section}' cannot appear after 'Additional Information' but before 'Comments'.")
                                previous_section_type = "user-defined"
                    
                        # Output any errors
                        if section_errors:
                            for error in section_errors:
                                errors.append(error)
            
                    # Check the cell below "Property label"
                     elif term == "Property label":
                        column_below_label = sheet[term_index][2:]
                        invalid_label = [(i + 3, cell.value) for i, cell in enumerate(column_below_label) if not re.match(r'.*', str(cell.value))]
                        if invalid_label:
                            invalid_rows = [str(row) for row, _ in invalid_label]
                            invalid_values = [str(value) for _, value in invalid_label]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Specify the property label as text format. Value(s) found: {', '.join(invalid_values)}")
                         # Dynamically find the "Section" column
                        if "Section" in second_row_values:
                            section_column_index = second_row_values.index("Section") + 1  # Find the index of the "Section" column
                            section_letter = index_to_excel_column(section_column_index)  # Convert index to Excel column letter
                            column_below_section = sheet[section_letter][2:]  # Get all cells below the "Section" header

                            # New check for "Notes" in "Property label" and "Additional Information" in "Section"
                            for i, cell in enumerate(column_below_label):
                                if cell.value == "Notes":
                                    section_value = column_below_section[i].value  # Get the value in the "Section" column for the same row
                                    if section_value != "Additional Information":
                                        errors.append(f"<strong>Error</strong>: 'Notes' found in the 'Property label' column at row {i + 5}, but corresponding 'Section' column does not contain 'Additional Information'. Value found: {section_value}")
                    
                    # Check the cell below "Data type"
                     elif term == "Data type":
                        column_below_type = sheet[term_index][2:]
                        invalid_type = [(i + 3, cell.value) for i, cell in enumerate(column_below_type) if cell.value not in ["INTEGER", "REAL", "VARCHAR", "MULTILINE_VARCHAR", "HYPERLINK", "BOOLEAN", "CONTROLLEDVOCABULARY", "XML", "TIMESTAMP", "DATE", "SAMPLE"]]
                        if invalid_type:
                            invalid_rows = [str(row) for row, _ in invalid_type]
                            invalid_values = [str(value) for _, value in invalid_type]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Accepted types: INTEGER, REAL, VARCHAR, MULTILINE_VARCHAR, HYPERLINK, BOOLEAN, CONTROLLEDVOCABULARY, XML, TIMESTAMP, DATE, SAMPLE. Value(s) found: {', '.join(invalid_values)}")
                    
                    # Check the column below "Vocabulary code"
                     elif term == "Vocabulary code":
                        column_below_vocab = sheet[term_index][2:]
                        invalid_vocab = [(i + 3, cell.value) for i, cell in enumerate(column_below_vocab) if cell.value is not None and not re.match(r'^\$?[A-Z0-9_.]+$', str(cell.value))]
                        if invalid_vocab:
                            invalid_rows = [str(row) for row, _ in invalid_vocab]
                            invalid_values = [str(value) for _, value in invalid_vocab]
                            errors.append(f"<strong>Error</strong>: Invalid vocabulary code found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Value(s) found: {', '.join(invalid_values)}")


    # Close the workbook after use
    workbook.close()
    if type(errors) == list:
        output = "\n\n⦿ ".join(errors)
    else:
        output = "".join(errors)
    if output == "":
        return "File content: OK!"
    else:
        return output
    

def search_entity(o, e_type, e_code):
    if e_type == "EXPERIMENT_TYPE":
        return o.get_collection_type(e_code)
        
    elif e_type == "SAMPLE_TYPE":
        return o.get_object_type(e_code)
    
    elif e_type == "DATASET_TYPE":
        return o.get_dataset_type(e_code)
    
    elif e_type == "VOCABULARY_TYPE":
        return o.get_vocabulary(e_code)
    
def get_entity_list(o, entity_type):
    if entity_type == "EXPERIMENT_TYPE":
        return o.get_collection_types()
    
    elif entity_type == "SAMPLE_TYPE":
        return o.get_object_types()
    
    elif entity_type == "DATASET_TYPE":
        return o.get_dataset_types()
    
    elif entity_type == "VOCABULARY_TYPE":
        return o.get_vocabularies()
    
def compare_objects(obj1, obj2):
    # Check if both are None or both are empty strings
    if (obj1 is None and obj2 == "") or (obj1 == "" and obj2 is None):
        return True
    elif (obj1 == "False" and obj2 == "FALSE") or (obj1 == "FALSE" and obj2 == "False"):
        return True
    elif (obj1 == "True" and obj2 == "TRUE") or (obj1 == "TRUE" and obj2 == "True"):
        return True
    else:
        return obj1 == obj2
    
def get_df_value(df, prop, attr):
    column_name = 'propertyType'
    
    # Check if 'propertyType' column exists in the DataFrame
    if column_name not in df.columns:
        return None  # or handle this case appropriately

    value_to_find = prop

    # Create a boolean mask for rows where the condition is met
    mask = df[column_name] == value_to_find

    # Use the boolean mask to filter the DataFrame
    filtered_df = df[mask]

    if not filtered_df.empty:
        return filtered_df[attr].iloc[0] if attr in filtered_df.columns else None
    else:
        return None
    


def check_entity_same_code(file_path, o, openbis_entity):
    errors = []
    description = ""
    auto_code = ""
    val_script = ""
    prefix_code = ""

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    entity_type = sheet['A1'].value
    second_row_values = [cell.value for cell in sheet[2]]
    
    for term in second_row_values:
        term_index = second_row_values.index(term)
        if term == "Code":
            entity_code = sheet.cell(row=3, column=term_index + 1).value
        elif term == "Description":
            description = sheet.cell(row=3, column=term_index + 1).value
        elif term == "Auto generate codes":
            auto_code = sheet.cell(row=3, column=term_index + 1).value
        elif term == "Validation script":
            val_script = sheet.cell(row=3, column=term_index + 1).value
        elif term == "Generated code prefix":
            prefix_code = sheet.cell(row=3, column=term_index + 1).value
            
        #format values to match
        if (auto_code == True): auto_code = "TRUE"
        if (auto_code == False): auto_code = "FALSE"
        if (val_script == None): val_script = ""
    
    
    openbis_description = openbis_entity.description
    openbis_auto_code = openbis_entity.autoGeneratedCode
    openbis_val_script = openbis_entity.validationPlugin
    openbis_prefix_code = openbis_entity.generatedCodePrefix
        
    #cast values to STRING t match Excel data
    if (openbis_auto_code == True): openbis_auto_code = "TRUE"
    if (openbis_auto_code == False): openbis_auto_code = "FALSE"
    if (openbis_val_script == None): openbis_val_script = ""
        
    #check description
    if (description != openbis_description):
        errors.append(f"The Description of ('{entity_type}') '{entity_code}' has been changed compared to the previous version.")
        description_pattern = re.compile(r".*//.*")
        if not description_pattern.match(description):
            errors.append("<strong>ERROR</strong>: The Description of ('{entity_type}') '{entity_code}' should follow the schema: English Description + '//' + German Description.")
        
    #check auto-generated codes
    if (auto_code != openbis_auto_code):
        errors.append(f"The value of “Auto generate codes” of ('{entity_type}') '{entity_code}' has been changed from '{openbis_auto_code}' to '{auto_code}'.")
    
    #check validation scripts
    if (val_script == "" and openbis_val_script != ""):
        errors.append(f"The validation script '{openbis_val_script}' has been deleted from ('{entity_type}') '{entity_code}'.")
    elif (val_script != "" and openbis_val_script == ""):
        errors.append(f"A validation script '{val_script}' has been added to ('{entity_type}') '{entity_code}'.")
    elif (val_script != openbis_val_script):
        errors.append(f"The validation script of ('{entity_type}') '{entity_code}' has been changed from '{openbis_val_script}' to '{val_script}'.")
        
    #check generated code prefix
    if (prefix_code != openbis_prefix_code):
        errors.append(f"The Code Prefix of ('{entity_type}') '{entity_code}' has been changed from '{openbis_prefix_code}' to '{prefix_code}'.")
            
            
    #get assigned properties from the excel file
    prop_headers = [cell.value for cell in sheet[4]]
    entity_properties = []
    term_index = prop_headers.index("Code") + 1
    term_letter = index_to_excel_column(term_index)
        
    for cell in sheet[term_letter][4:]:
        if cell.value is not None:
            entity_properties.append(cell.value)
        
    #get assigned properties from the openbis instance
    openbis_entity_properties = []
    for prop in openbis_entity.get_property_assignments():
        openbis_entity_properties.append(prop.permId)
            
    # Remove None values from both lists before sorting
    entity_properties = [prop for prop in entity_properties if prop is not None]
    openbis_entity_properties = [prop for prop in openbis_entity_properties if prop is not None]

    #check if the properties lists are the same
    if sorted(entity_properties) != sorted(openbis_entity_properties):
        errors.append(f"⦿ The set of Property Types assigned to the ('{entity_type}') '{entity_code}' has been changed compared to the previous version.")

            
    #check which properties has been added and removed
    deleted_properties = []
    added_properties = []
        
    deleted_properties = list(set(openbis_entity_properties) - set(entity_properties))
    added_properties = list(set(entity_properties) - set(openbis_entity_properties))
        
    for d_prop in deleted_properties:
        errors.append(f"The Property type assignment '{d_prop}' has been removed.")
    for a_prop in added_properties:
        errors.append(f"The Property type assignment '{a_prop}' has been added.")
            
            
# =============================================================================
#         for prop in openbis_entity.get_property_assignments():
#             print(prop.attrs.all())
# NOT WORKING ANYWAY; TO GET ALL THE REQUESTED VALUES
# =============================================================================

    #save dict with all the properties values from the entity in the instance
    openbis_properties_data = {}
    for prop in openbis_entity.get_property_assignments():
        openbis_properties_data[prop.code] = {
            "label": prop.label,
            "description": prop.description,
            "dataType": prop.dataType,
            "vocabulary": prop.vocabulary if prop.vocabulary is not None else "",
            "metaData" : prop.metaData
        }
            
    #save dict with all the properties values from the excel metadata file
    prop_headers = [cell.value for cell in sheet[4]]
    properties_data = {}
    term_index = prop_headers.index("Code") + 1
    term_letter = index_to_excel_column(term_index)
        

    for row in sheet.iter_rows(min_row=5, values_only=True):
        code_value = row[term_index - 1]  # Index is 0-based
        if code_value is not None:

            properties_data[code_value] = {
                'label': row[prop_headers.index('Property label')],
                'description': row[prop_headers.index('Description')],
                'dataType': row[prop_headers.index('Data type')],
                "vocabulary": row[prop_headers.index('Vocabulary code')] if row[prop_headers.index('Vocabulary code')] is not None else "",
                'metaData': {} if row[prop_headers.index('Metadata')] in (None, "") else row[prop_headers.index('Metadata')],
                'mandatory': row[prop_headers.index('Mandatory')],
                'section': row[prop_headers.index('Section')],
                'plugin': row[prop_headers.index('Dynamic script')],
                }
            
# =============================================================================
# COMPARE EXCEL PROPERTIES WITH ALL THE INSTANCE PROPERTIES, NOT ASSIGNED ONES
#     for key in properties_data.keys():
#         try:
#             prop_ob = o.get_property_type(key)
#             if not compare_objects(properties_data[key]['label'],prop_ob.label):
#                 errors.append(f"The label of Property type {key} has been changed compared to the previous version from {prop_ob.label} to {properties_data[key]['label']}.")
#             elif not compare_objects(properties_data[key]['description'],prop_ob.description):
#                 errors.append(f"The description of Property type {key} has been changed compared to the previous version from {prop_ob.description} to {properties_data[key]['description']}.")
#             elif not compare_objects(properties_data[key]['dataType'],prop_ob.dataType):
#                 errors.append(f"The data type of Property type {key} has been changed compared to the previous version from {prop_ob.dataType} to {properties_data[key]['dataType']}. This is only permissible for some cases, e.g., 'CONTROLLEDVOCABULARY' to 'VARCHAR'!")
#             elif not compare_objects(properties_data[key]['vocabulary'],prop_ob.vocabulary):
#                 errors.append(f"The vocabulary code of Property type {key} has been changed compared to the previous version from {prop_ob.vocabulary} to {properties_data[key]['vocabulary']}. This is not allowed.")
#             elif not compare_objects(properties_data[key]['metaData'],prop_ob.metaData):
#                 errors.append(f"The metadata of Property type {key} has been changed compared to the previous version from {prop_ob.metaData} to {properties_data[key]['metaData']}. This is not allowed.")
#         except ValueError:
#             continue
# =============================================================================
        

    assigned_properties = openbis_entity.get_property_assignments().df
    #properties present in the excel but not in openbis: not assigned
    not_assigned_properties =  set(properties_data.keys()) - set(openbis_properties_data.keys())
    
    #compare both dicts with sets of properties to check the differences
    for key in openbis_properties_data.keys() & properties_data.keys():
        for assigned_field in ["mandatory", "section", "plugin"]:
            excel_assigned = properties_data[key][assigned_field]
            openbis_assigned = get_df_value(assigned_properties, key, assigned_field)
            if not compare_objects(excel_assigned,openbis_assigned):
                if assigned_field == "mandatory":
                    if (str(openbis_assigned).upper() == "FALSE" and str(excel_assigned).upper() == "TRUE"):
                        errors.append(f"The value of the attribute 'Mandatory' of Property type {key} has been changed compared to the previous version from FALSE to TRUE.")
                    elif (str(openbis_assigned).upper() == "TRUE" and str(excel_assigned).upper() == "FALSE"):
                        errors.append(f"<strong>ERROR</strong>: The value of the attribute 'Mandatory' of Property type {key} has been changed compared to the previous version from TRUE to FALSE. This is NOT allowed")
                elif assigned_field == "section":
                    errors.append(f"The section of Property type {key} has been changed compared to the previous version from {openbis_assigned} to {excel_assigned}.")
                elif assigned_field == "plugin":
                    if (openbis_assigned == "" or openbis_assigned == None) and (excel_assigned != "" or excel_assigned != None):
                        errors.append(f"<em>WARNING</em>: A dynamic property script ({excel_assigned}) has been added retrospectively to the Property type {key}.")
                    elif (str(openbis_assigned).upper() != str(excel_assigned).upper()):
                        errors.append(f"<strong>ERROR</strong>: The dynamic property script of Property type {key} has been changed or deleted compared to the previous version. This is NOT allowed")
                   
        for field in ["label", "description", "dataType", "vocabulary", "metaData"]:
            value1 = openbis_properties_data[key][field]
            value2 = properties_data[key][field]
            if not compare_objects(value1,value2):
                if field == "label":
                    errors.append(f"The label of Property type {key} has been changed compared to the previous version from {value1} to {value2}.")
                elif field == "description":
                    errors.append(f"The description of Property type {key} has been changed compared to the previous version from {value1} to {value2}.")
                elif field == "dataType":
                    errors.append(f"<em>WARNING</em>: The data type of Property type {key} has been changed compared to the previous version from from {value1} to {value2}. This is only permissible for some cases, e.g., 'CONTROLLEDVOCABULARY' to 'VARCHAR'!")
                elif field == "vocabulary":
                    errors.append(f"<strong>ERROR</strong>: The vocabulary code of Property type {key} has been changed compared to the previous version from from {value1} to {value2}. This is not allowed.")
                elif field == "metaData":
                    errors.append(f"<strong>ERROR</strong>: The metadata of Property type {key} has been changed compared to the previous version from from {value1} to {value2}. This is not allowed.")


    for key in not_assigned_properties:
        try:
             prop_ob = o.get_property_type(key)
             if not compare_objects(properties_data[key]['label'],prop_ob.label):
                 errors.append(f"The label of Property type {key} has been changed compared to the previous version from {prop_ob.label} to {properties_data[key]['label']}.")
             elif not compare_objects(properties_data[key]['description'],prop_ob.description):
                 errors.append(f"The description of Property type {key} has been changed compared to the previous version from {prop_ob.description} to {properties_data[key]['description']}.")
             elif not compare_objects(properties_data[key]['dataType'],prop_ob.dataType):
                 errors.append(f"The data type of Property type {key} has been changed compared to the previous version from {prop_ob.dataType} to {properties_data[key]['dataType']}. This is only permissible for some cases, e.g., 'CONTROLLEDVOCABULARY' to 'VARCHAR'!")
             elif not compare_objects(properties_data[key]['vocabulary'],prop_ob.vocabulary):
                 errors.append(f"The vocabulary code of Property type {key} has been changed compared to the previous version from {prop_ob.vocabulary} to {properties_data[key]['vocabulary']}. This is not allowed.")
             elif not compare_objects(properties_data[key]['metaData'],prop_ob.metaData):
                 errors.append(f"The metadata of Property type {key} has been changed compared to the previous version from {prop_ob.metaData} to {properties_data[key]['metaData']}. This is not allowed.")
        except ValueError:
             continue
        
    workbook.close()
    
    return "\n\n⦿ ".join(errors)
        
def check_entity_diff_code(file_path, o):
    errors = []
    
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    entity_type = sheet['A1'].value

    if(entity_type) == "VOCABULARY_TYPE":
        return "\n".join(errors)
    
    openbis_entity_types = get_entity_list(o, entity_type)
    
    openbis_entity_properties = {}
    
    #get all the properties for each entity type from the instance, and save them in a dictionary
    for etype in openbis_entity_types:
        props_by_type = []
        openbis_entity_properties[etype.code] = []
        if etype.code != "UNKNOWN":
            for prop in etype.get_property_assignments():
                props_by_type.append(prop.permId)
            openbis_entity_properties[etype.code] = props_by_type
    
    #get the assigned properties of the entity in the excel
    entity_headers = [cell.value for cell in sheet[2]]
    entity_properties = []
    term_index = entity_headers.index("Code") + 1
    entity_code = sheet.cell(row=3, column=term_index).value
    term_letter = index_to_excel_column(term_index)
    
    for cell in sheet[term_letter][4:]:
        if cell.value is not None:
            entity_properties.append(cell.value)
            
    for key, prop_list in openbis_entity_properties.items():
        if set(prop_list) == set(entity_properties):
            errors.append(f"The {entity_type} '{entity_code}' is very similar to the existing {entity_type} '{key}'. Please consider whether you need to create a new entity type or whether you can re-use '{key}'")
    
    return "\n\n⦿ ".join(errors)


def check_prefix_sufix(file_path, o):
    errors = []
    
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    entity_type = sheet['A1'].value

    if(entity_type) == "VOCABULARY_TYPE":
        return "\n\n⦿ ".join(errors)
    
    entity_headers = [cell.value for cell in sheet[2]]
    term_index = entity_headers.index("Code") + 1
    entity_code = sheet.cell(row=3, column=term_index).value
    
    pattern = re.compile(r'^[A-Za-z0-9_.]+\.[A-Za-z0-9_]+$')
    
    if pattern.match(entity_code):
        parts = entity_code.rsplit('.', 1)
        prefix = parts[0]
        
        prop_headers = [cell.value for cell in sheet[4]]
        entity_properties = []
        term_index = prop_headers.index("Code") + 1
        term_letter = index_to_excel_column(term_index)
        
        for cell in sheet[term_letter][4:]:
            if cell.value is not None:
                entity_properties.append(cell.value)
        
        #get assigned properties from the openbis instance
        try:
            prefix_entity = search_entity(o, entity_type, prefix)
        except ValueError as e:
            errors.append(f"⦿ Entity type '{prefix}' is not present in the system, and cannot be the prefix of a new entity to be registered.")
            return "\n\n⦿ ".join(errors)
        
        prefix_properties = []
        for prop in prefix_entity.get_property_assignments():
            prefix_properties.append(prop.permId)
            
        #get the properties that are in the PREFIX but not in the SUFIX
        difference = [value for value in prefix_properties if value not in entity_properties]
        
        prefix_properties_data = {}
        for prop in prefix_entity.get_property_assignments():
            prefix_properties_data[prop.code] = {
                "label": prop.label,
                "description": prop.description,
                "dataType": prop.dataType,
                "vocabulary": prop.vocabulary if prop.vocabulary is not None else "",
                "metaData" : prop.metaData
            }
            
        entity_properties_data = {}
        for row in sheet.iter_rows(min_row=5, values_only=True):
            code_value = row[term_index - 1]  # Index is 0-based
            if code_value is not None:
                entity_properties_data[code_value] = {
                    'label': row[prop_headers.index('Property label')],
                    'description': row[prop_headers.index('Description')],
                    'dataType': row[prop_headers.index('Data type')],
                    "vocabulary": row[prop_headers.index('Vocabulary code')] if row[prop_headers.index('Vocabulary code')] is not None else "",
                    'metaData': {} if row[prop_headers.index('Metadata')] in (None, "") else row[prop_headers.index('Metadata')],
                    }

        changes = []
        #compare both dicts with sets of properties to check the differences
        for key in prefix_properties_data.keys() & entity_properties_data.keys():
            for field in ["label", "description", "dataType", "vocabulary", "metaData"]:
                value1 = prefix_properties_data[key][field]
                value2 = entity_properties_data[key][field]
                if value1 != value2:
                    if field == "label":
                        changes.append(f"Change in label of Property type {key}.")
                    elif field == "description":
                        changes.append(f"Change in description of Property type {key}.")
                    elif field == "dataType":
                        changes.append(f"Change in data type of Property type {key}.")
                    elif field == "vocabulary":
                        changes.append(f"Change in vocabulary code of Property type {key}.")
                    elif field == "metaData":
                        changes.append(f"Change in metadata of Property type {key}.")

        if (len(difference) != 0) or (len(changes) != 0):
            errors.append(f"As a specification of the entity type {prefix}, the entity type {entity_code} must include all Property types of {prefix} without any changes.")
            errors.append("The missing properties are: ")
            missing = ", ".join(difference)
            errors.append(missing)
            errors.append("\n")
            errors.append("The changed property attributes are: ")
            changed = "\n".join(changes)
            errors.append(changed)
    
            
        check_prefix_prefix(o, prefix, entity_type, errors)
    
    
    return "\n\n⦿ ".join(errors)


def check_prefix_prefix(o, prefix, entity_type, errors):
    if '.' in prefix:
        # Split the string by the last dot
        prefix_2, suffix = prefix.rsplit('.', 1)
        
        prefix_entity = search_entity(o, entity_type, prefix_2)
        suffix_entity = search_entity(o, entity_type, suffix)
        
        prefix_properties = []
        for prop in prefix_entity.get_property_assignments():
            prefix_properties.append(prop.permId)
            
        suffix_properties = []
        for prop in suffix_entity.get_property_assignments():
            suffix_properties.append(prop.permId)
            
        difference = [value for value in prefix_properties if value not in suffix_properties]
        
        prefix_properties_data = {}
        for prop in prefix_entity.get_property_assignments():
            prefix_properties_data[prop.code] = {
                "label": prop.label,
                "description": prop.description,
                "dataType": prop.dataType,
                "vocabulary": prop.vocabulary if prop.vocabulary is not None else "",
                "metaData" : prop.metaData
            }
            
        suffix_properties_data = {}
        for prop2 in suffix_entity.get_property_assignments():
            suffix_properties_data[prop2.code] = {
                "label": prop2.label,
                "description": prop2.description,
                "dataType": prop2.dataType,
                "vocabulary": prop2.vocabulary if prop2.vocabulary is not None else "",
                "metaData" : prop2.metaData
            }
            
        changes = []
        #compare both dicts with sets of properties to check the differences
        for key in prefix_properties_data.keys() & suffix_properties_data.keys():
            for field in ["label", "description", "dataType", "vocabulary", "metaData"]:
                value1 = prefix_properties_data[key][field]
                value2 = suffix_properties_data[key][field]
                if value1 != value2:
                    if field == "label":
                        changes.append(f"Change in label of Property type {key}.")
                    elif field == "description":
                        changes.append(f"Change in description of Property type {key}.")
                    elif field == "dataType":
                        changes.append(f"Change in data type of Property type {key}.")
                    elif field == "vocabulary":
                        changes.append(f"Change in vocabulary code of Property type {key}.")
                    elif field == "metaData":
                        changes.append(f"Change in metadata of Property type {key}.")

        if (len(difference) != 0) or (len(changes) != 0):
            errors.append(f"As a specification of the entity type {prefix}, the entity type {suffix} must include all Property types of {prefix} without any changes.")
            missing = ", ".join(difference)
            if missing != "":
                errors.append("The missing properties are: ")
                errors.append(missing)
            else:
                errors.append("There are no missing properties")
            errors.append("The changed property attributes are: ")
            changed = "\n\n⦿ ".join(changes)
            errors.append(changed)


        # Recursively call the function with the prefix
        check_prefix_prefix(o, prefix_2, entity_type, errors)
        
        
def entity_checker(file_path, o):
    
    errors = []
    
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    entity_type = sheet['A1'].value
    entity_headers = [cell.value for cell in sheet[2]]
    term_index = entity_headers.index("Code") + 1
    entity_code = sheet.cell(row=3, column=term_index).value
    
    try:
        openbis_entity = search_entity(o, entity_type, entity_code)
    except ValueError as e:
        errors.append(f"⦿ Entity type '{entity_code}' is a new entity type (not present in the system) to be registered.")
        openbis_entity = ""
    
        
    if (openbis_entity != ""):
        errors.append(f"⦿ Entity type '{entity_code}' already exists.")
        same_code_errors = check_entity_same_code(file_path, o, openbis_entity)
        errors.append(same_code_errors)
    else:
        diff_code_errors = check_entity_diff_code(file_path, o)
        errors.append(diff_code_errors)
        
    prefix_errors = check_prefix_sufix(file_path, o)
    errors.append(prefix_errors)
    
    
    return "\n\n".join(errors)

def generate_csv_and_download(o, instance):
    """
    Generates CSV in-memory and returns the rows and in-memory CSV content.
    """

    csv_file = io.StringIO()
    writer = csv.writer(csv_file)

    header = ["INSTANCE", "DATE"]
    current_date = datetime.now().strftime("%d-%m-%Y")
    info = [instance, current_date]

    # Fetch data from the server (using pybis) and serialize
    spaces = [space.code for space in o.get_spaces()]  # Convert Space objects to simple strings
    projects = [project.code for project in o.get_projects()]
    experiment_types = [exp.code for exp in o.get_experiment_types()]
    object_types = [obj.code for obj in o.get_object_types() if obj.code != "UNKNOWN"]
    material_types = [material.code for material in o.get_material_types()]
    dataset_types = [dataset.code for dataset in o.get_dataset_types()]
    vocabs = [vocab.code for vocab in o.get_vocabularies()]
    plugins = [plug.name for plug in o.get_plugins()]

    masterdata_headers = [
        f"SPACES ({len(spaces)})", f"PROJECTS ({len(projects)})", f"EXPERIMENT TYPES ({len(experiment_types)})",
        f"OBJECT TYPES ({len(object_types)})", f"DATASET TYPES ({len(dataset_types)})",
        f"VOCABULARIES ({len(vocabs)})", f"PLUGINS ({len(plugins)})", f"MATERIAL TYPES ({len(material_types)})"
    ]

    masterdata = [
        current_date,
        spaces,
        projects,
        experiment_types,
        object_types,
        dataset_types,
        vocabs,
        plugins,
        material_types
    ]

    csv_rows = []

    # Write headers
    writer.writerow(header)

    # Write instance info
    writer.writerow(info)
    csv_rows.append(info)

    # Write an empty row
    writer.writerow("")
    csv_rows.append("")

    # Write the master data headers
    writer.writerow(masterdata_headers)
    csv_rows.append(masterdata_headers)

    # Write master data rows
    max_length = max(len(data) for data in masterdata)
    for i in range(max_length):
        row = [data[i] if i < len(data) else "" for data in masterdata]
        writer.writerow(row)
        csv_rows.append(row)

     # Write empty row for separating sections
    writer.writerow("")
    csv_rows.append("")

    # Restoring the part that writes object properties by type
    writer.writerow(["PROPERTY LIST BY OBJECT TYPE"])
    csv_rows.append(["PROPERTY LIST BY OBJECT TYPE"])

    # Write object types
    writer.writerow(object_types)
    csv_rows.append(object_types)

    props_by_obj = []
    for obj in o.get_object_types():
        if obj.code == "UNKNOWN":
            continue
        props = []
        assignments_df = obj.get_property_assignments().df
        if 'propertyType' in assignments_df.columns:
            for prop in obj.get_property_assignments():
                props.append(f"{prop.code} ({str(prop.dataType).lower()})")
        props_by_obj.append(props)

    # Determine the maximum length of the object properties
    max_length_props = max(len(properties) for properties in props_by_obj)

    # Write object properties row by row
    for i in range(max_length_props):
        row = [prop_list[i] if i < len(prop_list) else "" for prop_list in props_by_obj]
        writer.writerow(row)
        csv_rows.append(row)

    # Prepare masterdata, including current date first
    masterdata_dict = {
        "current_date": current_date,
        "spaces": spaces,
        "projects": projects,
        "experiment_types": experiment_types,
        "object_types": object_types,
        "dataset_types": dataset_types,
        "vocabs": vocabs,
        "plugins": plugins,
        "material_types": material_types,
        "props_by_obj": props_by_obj
    }

    # Return CSV content and rows for display
    return csv_rows, csv_file.getvalue(), masterdata_dict